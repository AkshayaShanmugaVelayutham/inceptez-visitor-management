from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from pathlib import Path
import configparser

app = Flask(__name__)
CORS(app)

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///visitors.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ========================================
# SMART EXCEL FILE FINDER
# ========================================

CONFIG_FILE = 'config.ini'
DEFAULT_EXCEL_FILE = 'visitors_log.xlsx'

def get_excel_path():
    """Get Excel file path from config or search for it"""
    config = configparser.ConfigParser()
    
    # Read config file
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if 'FILE_LOCATION' in config:
            saved_path = config['FILE_LOCATION'].get('excel_path', DEFAULT_EXCEL_FILE)
            
            # Check if file exists at saved location
            if os.path.exists(saved_path):
                print(f"✅ Excel file found at: {saved_path}")
                return saved_path
            else:
                print(f"⚠️ Excel file not found at saved location: {saved_path}")
                print(f"🔍 Searching for Excel file...")
    
    # Search for Excel file in common locations
    search_locations = [
        DEFAULT_EXCEL_FILE,  # Current directory
        os.path.join(os.getcwd(), DEFAULT_EXCEL_FILE),
        os.path.join(Path.home(), 'Desktop', DEFAULT_EXCEL_FILE),
        os.path.join(Path.home(), 'Documents', DEFAULT_EXCEL_FILE),
        os.path.join(Path.home(), 'Downloads', DEFAULT_EXCEL_FILE),
        os.path.join(Path.home(), 'Documents', 'Inceptez_Visitors', DEFAULT_EXCEL_FILE),
    ]
    
    # Check each location
    for location in search_locations:
        if os.path.exists(location):
            print(f"✅ Excel file found at: {location}")
            save_excel_path(location)
            return location
    
    # If not found anywhere, create in backend folder
    print(f"📝 Creating new Excel file at: {DEFAULT_EXCEL_FILE}")
    save_excel_path(DEFAULT_EXCEL_FILE)
    return DEFAULT_EXCEL_FILE

def save_excel_path(path):
    """Save Excel file path to config"""
    config = configparser.ConfigParser()
    config['FILE_LOCATION'] = {
        'excel_path': path,
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    with open(CONFIG_FILE, 'w') as configfile:
        config.write(configfile)
    print(f"💾 Saved Excel file location to config: {path}")

def update_excel_path_if_moved():
    """Check if Excel file was moved and update config"""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if 'FILE_LOCATION' in config:
            saved_path = config['FILE_LOCATION'].get('excel_path', DEFAULT_EXCEL_FILE)
            
            # If file doesn't exist at saved location, search for it
            if not os.path.exists(saved_path):
                print(f"🔍 Excel file moved! Searching for new location...")
                return get_excel_path()
    
    return saved_path if 'saved_path' in locals() else get_excel_path()

# Get Excel file path
EXCEL_FILE = get_excel_path()

# ========================================
# Visitor Model
# ========================================

class Visitor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    purpose = db.Column(db.String(200), nullable=False)
    comments = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'phone': self.phone,
            'date': self.date,
            'purpose': self.purpose,
            'comments': self.comments,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }

# Initialize Excel file
def init_excel():
    global EXCEL_FILE
    EXCEL_FILE = update_excel_path_if_moved()  # Check if moved
    
    if not os.path.exists(EXCEL_FILE):
        # Create directory if it doesn't exist
        excel_dir = os.path.dirname(EXCEL_FILE)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors Log"
        
        # Headers
        headers = ['ID', 'Name', 'Phone Number', 'Date', 'Purpose of Visit', 'Comments']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel file created at: {EXCEL_FILE}")

# Add visitor to Excel
def add_to_excel(visitor):
    global EXCEL_FILE
    try:
        EXCEL_FILE = update_excel_path_if_moved()  # Check if moved before writing
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get next row number (subtract 1 for header to get ID)
        next_row = ws.max_row + 1
        visitor_id = next_row - 1  # First data row (row 2) gets ID 1
        
        # Prepare row data with sequential ID
        row_data = [
            visitor_id,
            visitor.name,
            visitor.phone,
            visitor.date,
            visitor.purpose,
            visitor.comments or ''
        ]
        
        # Add row
        ws.append(row_data)
        
        # Style the new row
        row_num = ws.max_row
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[row_num]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        
        wb.save(EXCEL_FILE)
        print(f"✅ Added visitor to Excel: Row {row_num}, ID {visitor_id} at {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"❌ Error adding to Excel: {e}")
        return False

# Rebuild Excel file with dynamic IDs
def rebuild_excel():
    global EXCEL_FILE
    try:
        EXCEL_FILE = update_excel_path_if_moved()  # Check if moved before rebuilding
        
        visitors = Visitor.query.order_by(Visitor.created_at).all()
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors Log"
        
        # Headers
        headers = ['ID', 'Name', 'Phone Number', 'Date', 'Purpose of Visit', 'Comments']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Add all visitors with sequential IDs
        for idx, visitor in enumerate(visitors, start=1):
            row_data = [
                idx,
                visitor.name,
                visitor.phone,
                visitor.date,
                visitor.purpose,
                visitor.comments or ''
            ]
            ws.append(row_data)
            
            # Style the row
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        
        wb.save(EXCEL_FILE)
        print(f"✅ Rebuilt Excel file with {len(visitors)} visitors at {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"❌ Error rebuilding Excel: {e}")
        return False

# Routes
@app.route('/api/visitors', methods=['GET'])
def get_visitors():
    visitors = Visitor.query.order_by(Visitor.created_at.desc()).all()
    return jsonify([v.to_dict() for v in visitors])

@app.route('/api/visitors', methods=['POST'])
def add_visitor():
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['name', 'phone', 'date', 'purpose']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Create new visitor
        visitor = Visitor(
            name=data['name'],
            phone=data['phone'],
            date=data['date'],
            purpose=data['purpose'],
            comments=data.get('comments', '')
        )
        
        # Save to database
        db.session.add(visitor)
        db.session.commit()
        
        # Add to Excel
        add_to_excel(visitor)
        
        return jsonify({
            'message': 'Visitor added successfully',
            'visitor': visitor.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/visitors/<int:visitor_id>', methods=['DELETE'])
def delete_visitor(visitor_id):
    try:
        visitor = Visitor.query.get_or_404(visitor_id)
        db.session.delete(visitor)
        db.session.commit()
        
        # Rebuild Excel file with updated sequential IDs
        rebuild_excel()
        
        return jsonify({'message': 'Visitor deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-location', methods=['GET'])
def get_excel_location():
    """API endpoint to get current Excel file location"""
    global EXCEL_FILE
    EXCEL_FILE = update_excel_path_if_moved()
    return jsonify({
        'excel_path': os.path.abspath(EXCEL_FILE),
        'exists': os.path.exists(EXCEL_FILE)
    })

# Initialize database and Excel
with app.app_context():
    db.create_all()
    init_excel()
    print(f"\n{'='*60}")
    print(f"🚀 Visitor Management System Started")
    print(f"{'='*60}")
    print(f"📊 Excel File Location: {os.path.abspath(EXCEL_FILE)}")
    print(f"💾 Database Location: {os.path.abspath('visitors.db')}")
    print(f"🌐 Backend URL: http://localhost:5000")
    print(f"{'='*60}\n")

if __name__ == '__main__':
    app.run(debug=True, port=5000)
