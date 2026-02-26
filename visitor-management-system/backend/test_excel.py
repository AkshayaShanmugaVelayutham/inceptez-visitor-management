"""
TEST SCRIPT - Run this to check if Excel is working
"""
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Get Desktop path
desktop = os.path.join(Path.home(), 'Desktop')
excel_file = os.path.join(desktop, 'TEST_Visitors_Log.xlsx')

print("="*70)
print("🧪 TESTING EXCEL FILE CREATION")
print("="*70)

print(f"\n1️⃣ Your Desktop path: {desktop}")
print(f"   Desktop exists? {os.path.exists(desktop)}")

print(f"\n2️⃣ Excel file will be saved to:")
print(f"   {excel_file}")

print(f"\n3️⃣ Creating test Excel file...")

try:
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Add headers
    headers = ['ID', 'Name', 'Phone', 'Date', 'Purpose', 'Comments']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add sample data
    ws.append([1, 'Test Person 1', '1234567890', '2026-02-09', 'Testing', 'First test'])
    ws.append([2, 'Test Person 2', '0987654321', '2026-02-09', 'Testing', 'Second test'])
    
    # Save file
    wb.save(excel_file)
    
    print(f"✅ SUCCESS! Excel file created!")
    print(f"\n4️⃣ File saved to:")
    print(f"   {excel_file}")
    print(f"   File exists? {os.path.exists(excel_file)}")
    print(f"   File size: {os.path.getsize(excel_file)} bytes")
    
    print(f"\n{'='*70}")
    print(f"✅ TEST PASSED!")
    print(f"{'='*70}")
    print(f"\n🎯 GO TO YOUR DESKTOP AND LOOK FOR:")
    print(f"   TEST_Visitors_Log.xlsx")
    print(f"\n   Open it in Excel to verify!")
    print(f"{'='*70}\n")
    
except Exception as e:
    print(f"\n❌ ERROR: {e}")
    print(f"\nTrying alternative location (current directory)...")
    
    try:
        excel_file = 'TEST_Visitors_Log.xlsx'
        wb.save(excel_file)
        print(f"✅ Saved to: {os.path.abspath(excel_file)}")
        print(f"   Look in the backend folder!")
    except Exception as e2:
        print(f"❌ Still failed: {e2}")
        print(f"\nPlease check:")
        print(f"1. openpyxl is installed: pip install openpyxl")
        print(f"2. You have write permissions")
