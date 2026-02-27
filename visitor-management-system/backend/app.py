from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from pathlib import Path
from config import EXCEL_FILE_PATH

import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app, origins=["*"])  # Allow all origins for development
import os

db_host = os.getenv("DB_HOST")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_name = os.getenv("DB_NAME")

# ============================================================
# DATABASE CONFIGURATION - MYSQL (SQL DATABASE)
# ============================================================

# Get database URL from environment variable
DATABASE_URL = os.environ.get('DATABASE_URL')

if not DATABASE_URL:
    print("\n" + "="*70)
    print("⚠️  WARNING: DATABASE_URL not found!")
    print("="*70)
    print("Please create backend/.env file with:")
    print("DATABASE_URL=mysql+pymysql://root:YOUR_PASSWORD@localhost:3306/visitors")
    print("="*70 + "\n")
    exit(1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ECHO'] = True  # Print all SQL queries (for demonstration)

db = SQLAlchemy(app)

print("\n" + "="*70)
print("🐬 CONNECTED TO MYSQL DATABASE (SQL)")
print("="*70)
print(f"Database: MySQL Server at localhost:3306/visitors")
print("="*70 + "\n")

# Use Excel path from config.py
EXCEL_FILE = EXCEL_FILE_PATH

# Visitor Model
class Visitor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    purpose = db.Column(db.String(200), nullable=False)
    meets_whom = db.Column(db.String(100), nullable=False)
    comments = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'phone': self.phone,
            'email': self.email,
            'date': self.date,
            'purpose': self.purpose,
            'meets_whom': self.meets_whom,
            'comments': self.comments,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }

# Initialize Excel file
def init_excel():
    try:
        # Make sure directory exists
        excel_dir = os.path.dirname(EXCEL_FILE)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir, exist_ok=True)
            print(f"Created directory: {excel_dir}")
        
        if not os.path.exists(EXCEL_FILE):
            print(f"Creating new Excel file: {EXCEL_FILE}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Visitors Log"
            
            # Headers - now includes Email ID, Meets Whom, and Created At
            headers = ['ID', 'Name', 'Phone Number', 'Email ID', 'Date', 'Purpose of Visit', 'Meets Whom', 'Comments', 'Created At']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 40
            ws.column_dimensions['I'].width = 20
            
            wb.save(EXCEL_FILE)
            print(f"✅ Excel file created successfully!")
        else:
            print(f"Excel file already exists: {EXCEL_FILE}")
    except Exception as e:
        print(f"❌ ERROR creating Excel file: {e}")
        print(f"📁 Attempted location: {EXCEL_FILE}")
        print(f"💡 TIP: Check folder permissions or change EXCEL_FILE location in app.py")

# Add visitor to Excel
def add_to_excel(visitor):
    try:
        # Check if Excel file exists, if not create it
        if not os.path.exists(EXCEL_FILE):
            print(f"📝 Excel file doesn't exist. Creating it now...")
            init_excel()
        
        # Load the Excel file
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get next row number (subtract 1 for header to get ID)
        next_row = ws.max_row + 1
        visitor_id = next_row - 1  # First data row (row 2) gets ID 1
        
        # Prepare row data with sequential ID
        row_data = [
            visitor_id,
            visitor.name,
            visitor.phone,
            visitor.date,
            visitor.purpose,
            visitor.comments or ''
        ]
        
        # Add row
        ws.append(row_data)
        
        # Style the new row
        row_num = ws.max_row
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[row_num]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        
        wb.save(EXCEL_FILE)
        print(f"✅ Added visitor to Excel: Row {row_num}, ID {visitor_id}")
        return True
    except PermissionError:
        print(f"❌ ERROR: Excel file is open in another program!")
        print(f"💡 TIP: Close Excel and try again")
        return False
    except Exception as e:
        print(f"❌ ERROR adding to Excel: {e}")
        print(f"📁 File location: {EXCEL_FILE}")
        return False

# Rebuild Excel file with dynamic IDs - GUARANTEED SYNCHRONIZATION
def rebuild_excel():
    try:
        print(f"\n{'='*70}")
        print(f"🔄 REBUILDING EXCEL FILE WITH NEW IDs")
        print(f"{'='*70}")
        
        # Get visitors in ORDER THEY WERE ENTERED (oldest first)
        visitors = Visitor.query.order_by(Visitor.created_at.asc()).all()
        print(f"\n📊 Database has {len(visitors)} visitors total")
        print(f"📋 Order: OLDEST FIRST (entry order)")
        
        # Delete old Excel file to ensure fresh start
        if os.path.exists(EXCEL_FILE):
            try:
                os.remove(EXCEL_FILE)
                print(f"🗑️ Deleted old Excel file")
            except:
                pass
        
        # Create brand new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors Log"
        
        # Headers - includes Email ID and Meets Whom
        headers = ['ID', 'Name', 'Phone Number', 'Email ID', 'Date', 'Purpose of Visit', 'Meets Whom', 'Comments']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        print(f"\n✏️ Adding visitors to Excel (OLDEST FIRST - entry order):")
        print(f"{'-'*70}")
        
        # Add all visitors with sequential IDs starting from 1
        for display_id, visitor in enumerate(visitors, start=1):
            row_data = [
                display_id,
                visitor.name,
                visitor.phone,
                visitor.email,
                visitor.date,
                visitor.purpose,
                visitor.meets_whom,
                visitor.comments or '',
                visitor.created_at.strftime('%Y-%m-%d %H:%M:%S')  # Format: 2026-02-17 10:30:45
            ]
            ws.append(row_data)
            
            # Print what we're adding
            print(f"   Excel Row {display_id}: {visitor.name} (DB ID: {visitor.id})")
            
            # Style the row
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 20
        
        # Try to save the file
        try:
            wb.save(EXCEL_FILE)
            
            print(f"{'-'*70}")
            print(f"✅ Excel file saved successfully!")
            print(f"📁 Location: {EXCEL_FILE}")
            print(f"📊 Total rows: {len(visitors)} visitors")
            print(f"🔢 IDs in Excel: 1, 2, 3... up to {len(visitors)}")
            print(f"✅ IDs MATCH BROWSER EXACTLY!")
            print(f"{'='*70}\n")
            return True
            
        except PermissionError:
            # File is locked - WPS Office/Excel has it open
            print(f"{'-'*70}")
            print(f"")
            print(f"🔒 ⚠️  EXCEL FILE IS LOCKED! ⚠️")
            print(f"")
            print(f"The Excel file is currently OPEN in WPS Office or Excel.")
            print(f"Windows won't let me update it while it's open.")
            print(f"")
            print(f"✅ Your data IS SAFE in the database!")
            print(f"")
            print(f"📝 TO UPDATE EXCEL:")
            print(f"   1. CLOSE WPS Office/Excel (completely quit the program)")
            print(f"   2. Run this command: python find_excel.py")
            print(f"   3. It will update and reopen with new data")
            print(f"")
            print(f"💡 TIP: Keep Excel CLOSED while adding/deleting visitors")
            print(f"        Open it only when you want to VIEW the final data")
            print(f"")
            print(f"📁 Excel location: {EXCEL_FILE}")
            print(f"{'='*70}\n")
            return False
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"❌ ERROR REBUILDING EXCEL")
        print(f"{'='*70}")
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        print(f"{'='*70}\n")
        return False

# Routes
@app.route('/api/visitors', methods=['GET'])
def get_visitors():
    # Show oldest first (entry order) - matches Excel
    visitors = Visitor.query.order_by(Visitor.created_at.asc()).all()
    return jsonify([v.to_dict() for v in visitors])

@app.route('/api/visitors', methods=['POST'])
def add_visitor():
    try:
        data = request.json
        print(f"\n➕ NEW VISITOR SUBMITTED:")
        print(f"   Name: {data.get('name')}")
        
        # Validate required fields
        required_fields = ['name', 'phone', 'email', 'date', 'purpose', 'meets_whom']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Create new visitor
        visitor = Visitor(
            name=data['name'],
            phone=data['phone'],
            email=data['email'],
            date=data['date'],
            purpose=data['purpose'],
            meets_whom=data['meets_whom'],
            comments=data.get('comments', '')
        )
        
        # Save to database
        db.session.add(visitor)
        db.session.commit()
        print(f"✅ Saved to database (DB ID: {visitor.id})")
        
        # Rebuild Excel to keep IDs synchronized
        print(f"🔄 Rebuilding Excel with updated IDs...")
        rebuild_excel()
        
        return jsonify({
            'message': 'Visitor added successfully',
            'visitor': visitor.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/visitors/<int:visitor_id>', methods=['DELETE'])
def delete_visitor(visitor_id):
    try:
        visitor = Visitor.query.get_or_404(visitor_id)
        print(f"\n🗑️ DELETING VISITOR:")
        print(f"   DB ID: {visitor.id}, Name: {visitor.name}")
        
        db.session.delete(visitor)
        db.session.commit()
        print(f"✅ Deleted from database")
        
        # Rebuild Excel file with updated sequential IDs
        print(f"🔄 Rebuilding Excel with updated IDs...")
        rebuild_excel()
        
        return jsonify({'message': 'Visitor deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR: {e}")
        return jsonify({'error': str(e)}), 500

# Initialize database and Excel
with app.app_context():
    db.create_all()
    init_excel()

if __name__ == '__main__':
    app.run(debug=True, port=5000)

app.run(host="0.0.0.0", port=5000)
